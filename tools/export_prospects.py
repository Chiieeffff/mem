"""
Export a list of iGaming prospects to XLSX and append them to suggested_companies.csv
so they are excluded from future searches.

Usage:
    python3 tools/export_prospects.py --input '[{"name":"Acme","website":"acme.com","linkedin":"linkedin.com/company/acme","region":"EU","why_fits":"White-label platform provider","size":"80","crypto_status":"No public crypto offering found"}]'

Output:
    .tmp/prospects_YYYYMMDD_HHMMSS.xlsx  (path printed to stdout)
    suggested_companies.csv              (auto-appended)
"""

import argparse
import csv
import json
import os
import sys
from datetime import datetime

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ── colours ──────────────────────────────────────────────────────────────────
NAVY  = "0D1B2A"
TEAL  = "1A7A6A"
MINT  = "E8F5F2"
WHITE = "FFFFFF"
BLACK = "1A1A1A"
LGREY = "F4F4F4"
MGREY = "CCCCCC"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color=BLACK, size=10):
    return Font(bold=bold, color=color, size=size, name="Calibri")

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def border():
    s = Side(style="thin", color=MGREY)
    return Border(left=s, right=s, top=s, bottom=s)

# ── build xlsx ────────────────────────────────────────────────────────────────
COLUMNS = [
    ("Company",              28),
    ("Website",              28),
    ("LinkedIn",             36),
    ("Region",               18),
    ("Why fits ICP",         40),
    ("Est. size",            12),
    ("Crypto status",        34),
]

def build_xlsx(prospects: list[dict], output_path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Prospects"

    # title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
    title_cell = ws.cell(row=1, column=1, value="iGaming Platform Prospects — GatewayCrypto")
    title_cell.fill = fill(NAVY)
    title_cell.font = Font(bold=True, color=WHITE, size=13, name="Calibri")
    title_cell.alignment = align(h="center")
    ws.row_dimensions[1].height = 28

    # date row
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(COLUMNS))
    date_cell = ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    date_cell.fill = fill(TEAL)
    date_cell.font = Font(color=WHITE, size=9, italic=True, name="Calibri")
    date_cell.alignment = align(h="right")

    # header row
    for col_idx, (header, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.fill = fill(TEAL)
        cell.font = font(bold=True, color=WHITE, size=10)
        cell.alignment = align(h="center")
        cell.border = border()
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
    ws.row_dimensions[3].height = 20

    # data rows
    fields = ["name", "website", "linkedin", "region", "why_fits", "size", "crypto_status"]
    for row_idx, company in enumerate(prospects, start=4):
        bg = MINT if row_idx % 2 == 0 else WHITE
        for col_idx, field in enumerate(fields, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=company.get(field, ""))
            cell.fill = fill(bg)
            cell.font = font(size=10)
            cell.alignment = align(wrap=True)
            cell.border = border()
        ws.row_dimensions[row_idx].height = 18

    # freeze panes below header
    ws.freeze_panes = "A4"

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)

# ── append to suggested_companies.csv ────────────────────────────────────────
SUGGESTED_CSV = os.path.join(os.path.dirname(__file__), "..", "suggested_companies.csv")

def append_to_suggested(prospects: list[dict]) -> None:
    csv_path = os.path.abspath(SUGGESTED_CSV)
    file_exists = os.path.exists(csv_path)
    with open(csv_path, "a", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        if not file_exists:
            writer.writerow(["company_name", "date_added"])
        for company in prospects:
            writer.writerow([company.get("name", ""), datetime.now().strftime("%Y-%m-%d")])

# ── main ──────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True, help="JSON array of prospect objects")
    parser.add_argument("--output", default=None, help="Output XLSX path (optional)")
    args = parser.parse_args()

    try:
        prospects = json.loads(args.input)
    except json.JSONDecodeError as e:
        print(f"ERROR: Invalid JSON input — {e}", file=sys.stderr)
        sys.exit(1)

    if not prospects:
        print("ERROR: Empty prospect list", file=sys.stderr)
        sys.exit(1)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = args.output or os.path.abspath(f".tmp/prospects_{timestamp}.xlsx")

    build_xlsx(prospects, output_path)
    append_to_suggested(prospects)

    print(output_path)

if __name__ == "__main__":
    main()
