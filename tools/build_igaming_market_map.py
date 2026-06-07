"""
Build top 100 iGaming platforms market map with LinkedIn links.
Output: Company | Website | LinkedIn | Est. Size | Crypto Processor
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Full dataset — top 100 iGaming B2B platforms
# linkedin_slug is the path segment after linkedin.com/company/
COMPANIES = [
    {
        "Company": "EveryMatrix",
        "Website": "everymatrix.com",
        "LinkedIn": "linkedin.com/company/everymatrix",
        "Est. Size": "750+",
        "Crypto Processor": "MoneyMatrix (native)",
    },
    {
        "Company": "SOFTSWISS",
        "Website": "softswiss.com",
        "LinkedIn": "linkedin.com/company/softswiss",
        "Est. Size": "1,000+",
        "Crypto Processor": "Native Crypto, CoinsPaid",
    },
    {
        "Company": "BetConstruct",
        "Website": "betconstruct.com",
        "LinkedIn": "linkedin.com/company/betconstruct",
        "Est. Size": "5,000+",
        "Crypto Processor": "Native Crypto (BTC, ETH, USDT+)",
    },
    {
        "Company": "Gaming Innovation Group (GiG)",
        "Website": "gig.com",
        "LinkedIn": "linkedin.com/company/gaming-innovation-group",
        "Est. Size": "570+",
        "Crypto Processor": "—",
    },
    {
        "Company": "Kambi",
        "Website": "kambi.com",
        "LinkedIn": "linkedin.com/company/kambi",
        "Est. Size": "870+",
        "Crypto Processor": "—",
    },
    {
        "Company": "Slotegrator",
        "Website": "slotegrator.pro",
        "LinkedIn": "linkedin.com/company/slotegrator",
        "Est. Size": "150–280",
        "Crypto Processor": "Native Crypto, Moneygrator",
    },
    {
        "Company": "SoftGamings",
        "Website": "softgamings.com",
        "LinkedIn": "linkedin.com/company/softgamings",
        "Est. Size": "50–150",
        "Crypto Processor": "Native Crypto (BTC, ETH, LTC+)",
    },
    {
        "Company": "NuxGame",
        "Website": "nuxgame.com",
        "LinkedIn": "linkedin.com/company/nuxgame",
        "Est. Size": "50–150",
        "Crypto Processor": "Native Crypto (BTC, ETH, USDT)",
    },
    {
        "Company": "Digitain",
        "Website": "digitain.com",
        "LinkedIn": "linkedin.com/company/digitain",
        "Est. Size": "500+",
        "Crypto Processor": "—",
    },
    {
        "Company": "Pronet Gaming",
        "Website": "pronetgaming.com",
        "LinkedIn": "linkedin.com/company/pronet-gaming",
        "Est. Size": "200+",
        "Crypto Processor": "—",
    },
    {
        "Company": "WA.Technology",
        "Website": "watechnology.com",
        "LinkedIn": "linkedin.com/company/wa-technology",
        "Est. Size": "100–200",
        "Crypto Processor": "—",
    },
    {
        "Company": "Altenar",
        "Website": "altenar.com",
        "LinkedIn": "linkedin.com/company/altenar",
        "Est. Size": "400+",
        "Crypto Processor": "CoinsPaid",
    },
    {
        "Company": "BtoBet",
        "Website": "btobet.com",
        "LinkedIn": "linkedin.com/company/btobet",
        "Est. Size": "100–200",
        "Crypto Processor": "—",
    },
    {
        "Company": "BETBY",
        "Website": "betby.com",
        "LinkedIn": "linkedin.com/company/betby",
        "Est. Size": "150–200",
        "Crypto Processor": "—",
    },
    {
        "Company": "GR8 Tech",
        "Website": "gr8.tech",
        "LinkedIn": "linkedin.com/company/gr8-tech",
        "Est. Size": "500–600",
        "Crypto Processor": "Cryptopay",
    },
    {
        "Company": "Soft2Bet",
        "Website": "soft2bet.com",
        "LinkedIn": "linkedin.com/company/soft2bet",
        "Est. Size": "200–230",
        "Crypto Processor": "Native Crypto module",
    },
    {
        "Company": "iGP (iGaming Platform)",
        "Website": "igamingplatform.com",
        "LinkedIn": "linkedin.com/company/igaming-platform",
        "Est. Size": "100–200",
        "Crypto Processor": "Native Crypto + Fiat",
    },
    {
        "Company": "NSoft",
        "Website": "nsoft.com",
        "LinkedIn": "linkedin.com/company/nsoft",
        "Est. Size": "310+",
        "Crypto Processor": "—",
    },
    {
        "Company": "Pragmatic Solutions",
        "Website": "pragmatic.solutions",
        "LinkedIn": "linkedin.com/company/pragmaticsolutions",
        "Est. Size": "200–280",
        "Crypto Processor": "inabit (crypto add-on)",
    },
    {
        "Company": "Ultraplay",
        "Website": "ultraplay.com",
        "LinkedIn": "linkedin.com/company/ultraplay",
        "Est. Size": "100–200",
        "Crypto Processor": "—",
    },
    {
        "Company": "Sportingtech",
        "Website": "sportingtech.com",
        "LinkedIn": "linkedin.com/company/sportingtech",
        "Est. Size": "100–200",
        "Crypto Processor": "—",
    },
    {
        "Company": "BetStarters",
        "Website": "betstarters.com",
        "LinkedIn": "linkedin.com/company/betstarters",
        "Est. Size": "50–100",
        "Crypto Processor": "—",
    },
    {
        "Company": "Gamingtec",
        "Website": "gamingtec.com",
        "LinkedIn": "linkedin.com/company/gamingtec",
        "Est. Size": "50–150",
        "Crypto Processor": "—",
    },
    {
        "Company": "EvenBet Gaming",
        "Website": "evenbetgaming.com",
        "LinkedIn": "linkedin.com/company/evenbet-gaming",
        "Est. Size": "50–150",
        "Crypto Processor": "Native Crypto (poker solutions)",
    },
    {
        "Company": "GammaStack",
        "Website": "gammastack.com",
        "LinkedIn": "linkedin.com/company/gammastack",
        "Est. Size": "100–300",
        "Crypto Processor": "Native Crypto casino tools",
    },
    {
        "Company": "Atlaslive",
        "Website": "atlaslive.tech",
        "LinkedIn": "linkedin.com/company/atlaslive",
        "Est. Size": "50–150",
        "Crypto Processor": "—",
    },
    {
        "Company": "Salsa Technology",
        "Website": "salsatechnology.com",
        "LinkedIn": "linkedin.com/company/salsa-technology",
        "Est. Size": "100–200",
        "Crypto Processor": "—",
    },
    {
        "Company": "White Hat Gaming",
        "Website": "whitehatgaming.com",
        "LinkedIn": "linkedin.com/company/white-hat-gaming",
        "Est. Size": "100–200",
        "Crypto Processor": "—",
    },
    {
        "Company": "Bede Gaming",
        "Website": "bedegaming.com",
        "LinkedIn": "linkedin.com/company/bede-gaming",
        "Est. Size": "260+",
        "Crypto Processor": "—",
    },
    {
        "Company": "The Mill Adventure",
        "Website": "themilladventure.com",
        "LinkedIn": "linkedin.com/company/the-mill-adventure",
        "Est. Size": "30–80",
        "Crypto Processor": "—",
    },
    {
        "Company": "Infingame",
        "Website": "infingame.com",
        "LinkedIn": "linkedin.com/company/infingame",
        "Est. Size": "50–150",
        "Crypto Processor": "—",
    },
    {
        "Company": "Bragg Gaming",
        "Website": "bragg.games",
        "LinkedIn": "linkedin.com/company/bragg-gaming-group",
        "Est. Size": "100–300",
        "Crypto Processor": "—",
    },
    {
        "Company": "Uplatform",
        "Website": "uplatform.com",
        "LinkedIn": "linkedin.com/company/uplatform",
        "Est. Size": "300+",
        "Crypto Processor": "Integrated (500+ processors)",
    },
    {
        "Company": "Delasport",
        "Website": "delasport.com",
        "LinkedIn": "linkedin.com/company/delasport",
        "Est. Size": "425+",
        "Crypto Processor": "—",
    },
    {
        "Company": "Upgaming",
        "Website": "upgaming.com",
        "LinkedIn": "linkedin.com/company/upgaming",
        "Est. Size": "178+",
        "Crypto Processor": "Native Crypto + Fiat",
    },
    {
        "Company": "PlaylogiQ",
        "Website": "playlogiq.com",
        "LinkedIn": "linkedin.com/company/playlogiq",
        "Est. Size": "30–100",
        "Crypto Processor": "—",
    },
    {
        "Company": "GBO Gaming",
        "Website": "gbogaming.com",
        "LinkedIn": "linkedin.com/company/gbo-gaming",
        "Est. Size": "50–150",
        "Crypto Processor": "Native Crypto tools",
    },
    {
        "Company": "Hub88",
        "Website": "hub88.io",
        "LinkedIn": "linkedin.com/company/hub88",
        "Est. Size": "50+",
        "Crypto Processor": "—",
    },
    {
        "Company": "Pariplay",
        "Website": "pariplay.com",
        "LinkedIn": "linkedin.com/company/pariplay",
        "Est. Size": "50–130",
        "Crypto Processor": "—",
    },
    {
        "Company": "Groove Technologies",
        "Website": "groovetech.com",
        "LinkedIn": "linkedin.com/company/groove-technologies",
        "Est. Size": "50–150",
        "Crypto Processor": "Crypto-friendly tools",
    },
    {
        "Company": "Oddin.gg",
        "Website": "oddin.gg",
        "LinkedIn": "linkedin.com/company/oddin-gg",
        "Est. Size": "30–80",
        "Crypto Processor": "—",
    },
    {
        "Company": "GAN Sports",
        "Website": "gansports.com",
        "LinkedIn": "linkedin.com/company/gan-sports",
        "Est. Size": "200–400",
        "Crypto Processor": "—",
    },
    {
        "Company": "Amelco",
        "Website": "amelco.com",
        "LinkedIn": "linkedin.com/company/amelco",
        "Est. Size": "50–150",
        "Crypto Processor": "—",
    },
    {
        "Company": "Limeup",
        "Website": "limeup.io",
        "LinkedIn": "linkedin.com/company/limeup",
        "Est. Size": "100–200",
        "Crypto Processor": "Native Crypto casino",
    },
    {
        "Company": "Betinvest",
        "Website": "betinvest.com",
        "LinkedIn": "linkedin.com/company/betinvest",
        "Est. Size": "200+",
        "Crypto Processor": "—",
    },
    {
        "Company": "Kiron Interactive",
        "Website": "kiron-interactive.com",
        "LinkedIn": "linkedin.com/company/kiron-interactive",
        "Est. Size": "100–200",
        "Crypto Processor": "—",
    },
    {
        "Company": "Leap Gaming",
        "Website": "leapgaming.com",
        "LinkedIn": "linkedin.com/company/leap-gaming",
        "Est. Size": "50–100",
        "Crypto Processor": "—",
    },
    {
        "Company": "Scout Gaming",
        "Website": "scoutgaming.group",
        "LinkedIn": "linkedin.com/company/scout-gaming-group",
        "Est. Size": "50–100",
        "Crypto Processor": "—",
    },
    {
        "Company": "PopOk Gaming",
        "Website": "popokgaming.com",
        "LinkedIn": "linkedin.com/company/popok-gaming",
        "Est. Size": "100–200",
        "Crypto Processor": "—",
    },
    {
        "Company": "Ela Games",
        "Website": "elagames.com",
        "LinkedIn": "linkedin.com/company/ela-games",
        "Est. Size": "50–150",
        "Crypto Processor": "—",
    },
    {
        "Company": "ProgressPlay",
        "Website": "progressplay.com",
        "LinkedIn": "linkedin.com/company/progressplay",
        "Est. Size": "100–200",
        "Crypto Processor": "—",
    },
    {
        "Company": "Aspire Global",
        "Website": "aspireglobal.com",
        "LinkedIn": "linkedin.com/company/aspire-global",
        "Est. Size": "200–300",
        "Crypto Processor": "—",
    },
    {
        "Company": "CT Interactive",
        "Website": "ctinteractive.eu",
        "LinkedIn": "linkedin.com/company/ct-interactive",
        "Est. Size": "100–200",
        "Crypto Processor": "—",
    },
    {
        "Company": "Elbet",
        "Website": "elbet.eu",
        "LinkedIn": "linkedin.com/company/elbet",
        "Est. Size": "100–200",
        "Crypto Processor": "—",
    },
    {
        "Company": "Incentive Games",
        "Website": "incentivegames.com",
        "LinkedIn": "linkedin.com/company/incentive-games",
        "Est. Size": "50–150",
        "Crypto Processor": "—",
    },
    {
        "Company": "BetGames",
        "Website": "betgames.tv",
        "LinkedIn": "linkedin.com/company/betgames-tv",
        "Est. Size": "100–200",
        "Crypto Processor": "—",
    },
    {
        "Company": "Galaxsys",
        "Website": "galaxsys.com",
        "LinkedIn": "linkedin.com/company/galaxsys",
        "Est. Size": "50–150",
        "Crypto Processor": "—",
    },
    {
        "Company": "Spribe",
        "Website": "spribe.co",
        "LinkedIn": "linkedin.com/company/spribe",
        "Est. Size": "100–200",
        "Crypto Processor": "—",
    },
    {
        "Company": "LuckyStreak",
        "Website": "luckystreak.am",
        "LinkedIn": "linkedin.com/company/luckystreak",
        "Est. Size": "50–100",
        "Crypto Processor": "—",
    },
    {
        "Company": "Vivo Gaming",
        "Website": "vivogaming.com",
        "LinkedIn": "linkedin.com/company/vivo-gaming",
        "Est. Size": "50–150",
        "Crypto Processor": "—",
    },
    {
        "Company": "1x2 Network",
        "Website": "1x2network.com",
        "LinkedIn": "linkedin.com/company/1x2-network",
        "Est. Size": "50–150",
        "Crypto Processor": "—",
    },
    {
        "Company": "Iforium",
        "Website": "iforium.com",
        "LinkedIn": "linkedin.com/company/iforium",
        "Est. Size": "50–100",
        "Crypto Processor": "—",
    },
    {
        "Company": "BetSolutions",
        "Website": "betsolutions.am",
        "LinkedIn": "linkedin.com/company/betsolutions",
        "Est. Size": "50–100",
        "Crypto Processor": "Native Crypto (Zeppelin Crash)",
    },
    {
        "Company": "TG.Lab",
        "Website": "tglab.com",
        "LinkedIn": "linkedin.com/company/tg-lab",
        "Est. Size": "100–150",
        "Crypto Processor": "—",
    },
    {
        "Company": "Gamanza Group",
        "Website": "gamanza.com",
        "LinkedIn": "linkedin.com/company/gamanza",
        "Est. Size": "50–100",
        "Crypto Processor": "—",
    },
    {
        "Company": "Sportech",
        "Website": "sportech.com",
        "LinkedIn": "linkedin.com/company/sportech",
        "Est. Size": "100–200",
        "Crypto Processor": "—",
    },
    {
        "Company": "Sportnco",
        "Website": "sportnco.com",
        "LinkedIn": "linkedin.com/company/sportnco",
        "Est. Size": "100–200",
        "Crypto Processor": "—",
    },
    {
        "Company": "OpenBet",
        "Website": "openbet.com",
        "LinkedIn": "linkedin.com/company/openbet",
        "Est. Size": "500+",
        "Crypto Processor": "—",
    },
    {
        "Company": "Global Bet",
        "Website": "globalbet.com",
        "LinkedIn": "linkedin.com/company/global-bet",
        "Est. Size": "50–150",
        "Crypto Processor": "—",
    },
    {
        "Company": "BAAS iGaming",
        "Website": "baasigaming.com",
        "LinkedIn": "linkedin.com/company/baas-igaming",
        "Est. Size": "30–80",
        "Crypto Processor": "—",
    },
    {
        "Company": "Tecpinion",
        "Website": "tecpinion.com",
        "LinkedIn": "linkedin.com/company/tecpinion",
        "Est. Size": "50–150",
        "Crypto Processor": "—",
    },
    {
        "Company": "PieGaming",
        "Website": "piegaming.com",
        "LinkedIn": "linkedin.com/company/piegaming",
        "Est. Size": "30–80",
        "Crypto Processor": "—",
    },
    {
        "Company": "BetMakers Technology",
        "Website": "betmakers.com",
        "LinkedIn": "linkedin.com/company/betmakers-technology",
        "Est. Size": "100–200",
        "Crypto Processor": "—",
    },
    {
        "Company": "Vibra Gaming",
        "Website": "vibragaming.com",
        "LinkedIn": "linkedin.com/company/vibra-gaming",
        "Est. Size": "50–150",
        "Crypto Processor": "—",
    },
    {
        "Company": "Nucleus Gaming",
        "Website": "nucleusgaming.com",
        "LinkedIn": "linkedin.com/company/nucleus-gaming",
        "Est. Size": "30–80",
        "Crypto Processor": "—",
    },
    {
        "Company": "GreenTube",
        "Website": "greentube.com",
        "LinkedIn": "linkedin.com/company/greentube",
        "Est. Size": "100–300",
        "Crypto Processor": "—",
    },
    {
        "Company": "Singular",
        "Website": "singular.com",
        "LinkedIn": "linkedin.com/company/singular",
        "Est. Size": "50–150",
        "Crypto Processor": "—",
    },
    {
        "Company": "Betware",
        "Website": "betware.com",
        "LinkedIn": "linkedin.com/company/betware",
        "Est. Size": "50–150",
        "Crypto Processor": "—",
    },
    {
        "Company": "CT Gaming International",
        "Website": "ctgaminginternational.com",
        "LinkedIn": "linkedin.com/company/ct-gaming-international",
        "Est. Size": "100–200",
        "Crypto Processor": "—",
    },
    {
        "Company": "Inbet Games",
        "Website": "inbet.com",
        "LinkedIn": "linkedin.com/company/inbet-games",
        "Est. Size": "50–150",
        "Crypto Processor": "—",
    },
    {
        "Company": "Quadcode",
        "Website": "quadcode.com",
        "LinkedIn": "linkedin.com/company/quadcode",
        "Est. Size": "200–400",
        "Crypto Processor": "—",
    },
    {
        "Company": "Amusnet Interactive",
        "Website": "amusnet.com",
        "LinkedIn": "linkedin.com/company/amusnet-interactive",
        "Est. Size": "1,000+",
        "Crypto Processor": "—",
    },
    {
        "Company": "Stakelogic",
        "Website": "stakelogic.com",
        "LinkedIn": "linkedin.com/company/stakelogic",
        "Est. Size": "50–150",
        "Crypto Processor": "—",
    },
    {
        "Company": "Inplaysoft",
        "Website": "inplaysoft.com",
        "LinkedIn": "linkedin.com/company/inplaysoft",
        "Est. Size": "50–100",
        "Crypto Processor": "—",
    },
    {
        "Company": "AIS Technolabs",
        "Website": "aistechnolabs.com",
        "LinkedIn": "linkedin.com/company/ais-technolabs",
        "Est. Size": "200–500",
        "Crypto Processor": "Native Crypto casino",
    },
    {
        "Company": "Hero Gaming",
        "Website": "herogaming.com",
        "LinkedIn": "linkedin.com/company/hero-gaming",
        "Est. Size": "100–200",
        "Crypto Processor": "—",
    },
    {
        "Company": "7777 Gaming",
        "Website": "7777gaming.com",
        "LinkedIn": "linkedin.com/company/7777-gaming",
        "Est. Size": "50–150",
        "Crypto Processor": "—",
    },
    {
        "Company": "Authentic Gaming",
        "Website": "authenticgaming.com",
        "LinkedIn": "linkedin.com/company/authentic-gaming",
        "Est. Size": "50–100",
        "Crypto Processor": "—",
    },
    {
        "Company": "SkyWind Group",
        "Website": "skywindgroup.com",
        "LinkedIn": "linkedin.com/company/skywind-group",
        "Est. Size": "100–300",
        "Crypto Processor": "—",
    },
    {
        "Company": "Ezugi",
        "Website": "ezugi.com",
        "LinkedIn": "linkedin.com/company/ezugi",
        "Est. Size": "100–300",
        "Crypto Processor": "—",
    },
    {
        "Company": "BetinSport",
        "Website": "betinsport.com",
        "LinkedIn": "linkedin.com/company/betinsport",
        "Est. Size": "50–150",
        "Crypto Processor": "—",
    },
    {
        "Company": "Everi",
        "Website": "everi.com",
        "LinkedIn": "linkedin.com/company/everi",
        "Est. Size": "500+",
        "Crypto Processor": "—",
    },
    {
        "Company": "Softlab-NSK",
        "Website": "softlab-nsk.com",
        "LinkedIn": "linkedin.com/company/softlab-nsk",
        "Est. Size": "100–200",
        "Crypto Processor": "—",
    },
    {
        "Company": "Patagonia Entertainment",
        "Website": "patagoniaent.com",
        "LinkedIn": "linkedin.com/company/patagonia-entertainment",
        "Est. Size": "50–150",
        "Crypto Processor": "—",
    },
    {
        "Company": "Boomerang Partners",
        "Website": "boomerangpartners.com",
        "LinkedIn": "linkedin.com/company/boomerang-partners",
        "Est. Size": "50–150",
        "Crypto Processor": "—",
    },
    {
        "Company": "Rebet Technology",
        "Website": "rebet.com",
        "LinkedIn": "linkedin.com/company/rebet-technology",
        "Est. Size": "30–80",
        "Crypto Processor": "—",
    },
    {
        "Company": "Frently Solutions",
        "Website": "frently.com",
        "LinkedIn": "linkedin.com/company/frently",
        "Est. Size": "30–80",
        "Crypto Processor": "—",
    },
    {
        "Company": "Broadway Gaming",
        "Website": "broadwayplatform.com",
        "LinkedIn": "linkedin.com/company/broadway-gaming",
        "Est. Size": "30–80",
        "Crypto Processor": "—",
    },
    {
        "Company": "GammaSweep",
        "Website": "gammasweep.com",
        "LinkedIn": "linkedin.com/company/gammasweep",
        "Est. Size": "50–150",
        "Crypto Processor": "—",
    },
    {
        "Company": "GammaPlus",
        "Website": "gammaplus.io",
        "LinkedIn": "linkedin.com/company/gammaplus",
        "Est. Size": "50–150",
        "Crypto Processor": "—",
    },
]

assert len(COMPANIES) == 100, f"Expected 100, got {len(COMPANIES)}"

df = pd.DataFrame(COMPANIES)

# ── Excel output ──────────────────────────────────────────────────────────────
OUT = "/Users/roman/claude/.tmp/igaming_top100_market_map.xlsx"
wb = Workbook()
ws = wb.active
ws.title = "Top 100 iGaming Platforms"

COLS = ["Company", "Website", "LinkedIn", "Est. Size", "Crypto Processor"]
COL_WIDTHS = [32, 28, 42, 14, 32]

# Header style
HEADER_FILL = PatternFill("solid", fgColor="1A1A2E")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ALT_FILL = PatternFill("solid", fgColor="F2F4F7")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
CRYPTO_FILL = PatternFill("solid", fgColor="E8F5E9")  # light green for companies with crypto
BORDER = Border(
    bottom=Side(style="thin", color="D0D5DD"),
)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)

# Title row
ws.merge_cells("A1:E1")
title_cell = ws["A1"]
title_cell.value = f"Top 100 iGaming B2B Platforms — GatewayCrypto Market Map (2026-06-02)"
title_cell.font = Font(bold=True, size=13, color="1A1A2E")
title_cell.alignment = CENTER
ws.row_dimensions[1].height = 28

# Header row
for col_idx, col_name in enumerate(COLS, 1):
    cell = ws.cell(row=2, column=col_idx, value=col_name)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = BORDER
ws.row_dimensions[2].height = 22

# Data rows
for row_idx, rec in enumerate(COMPANIES, 3):
    fill = CRYPTO_FILL if rec["Crypto Processor"] != "—" else (ALT_FILL if row_idx % 2 == 0 else WHITE_FILL)
    for col_idx, col_name in enumerate(COLS, 1):
        val = rec[col_name]
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.fill = fill
        cell.alignment = LEFT
        cell.border = BORDER
        cell.font = Font(size=10)
    ws.row_dimensions[row_idx].height = 18

# Column widths
for col_idx, width in enumerate(COL_WIDTHS, 1):
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# Freeze header rows
ws.freeze_panes = "A3"

wb.save(OUT)
print(f"Saved: {OUT}")
print(f"Total rows: {len(COMPANIES)}")

# Also print a plain-text preview (tab-separated)
print("\n" + "\t".join(COLS))
print("-" * 120)
for rec in COMPANIES:
    print("\t".join(rec[c] for c in COLS))
