import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

NAVY  = "0D1B2A"
TEAL  = "008B8B"
MINT  = "E8F8F5"
LGREY = "F4F4F4"
AMBER = "FFF3CD"
WHITE = "FFFFFF"
BLACK = "1A1A1A"
GREEN = "1A7A4A"
RED   = "C0392B"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color=BLACK, size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def header_cell(ws, row, col, text, bg=NAVY, fg=WHITE, size=10, width=None):
    c = ws.cell(row=row, column=col, value=text)
    c.fill = fill(bg)
    c.font = font(bold=True, color=fg, size=size)
    c.alignment = align("center")
    c.border = thin_border()
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c

def data_cell(ws, row, col, text, bg=WHITE, bold=False, color=BLACK, wrap=True, h="left"):
    c = ws.cell(row=row, column=col, value=text)
    c.fill = fill(bg)
    c.font = font(bold=bold, color=color)
    c.alignment = align(h, wrap=wrap)
    c.border = thin_border()
    return c

def row_bg(i):
    return MINT if i % 2 == 0 else LGREY

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def freeze(ws, cell="A2"):
    ws.freeze_panes = cell


# ── Sheet 1: Overview ─────────────────────────────────────────────────────────

def build_overview(wb):
    ws = wb.active
    ws.title = "Overview"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [28, 55])
    ws.row_dimensions[1].height = 40

    # Title
    ws.merge_cells("A1:B1")
    c = ws["A1"]
    c.value = "next.io 2026 — Event Strategy Plan"
    c.fill = fill(NAVY)
    c.font = Font(bold=True, color=WHITE, size=18, name="Calibri")
    c.alignment = align("center")
    c.border = thin_border()

    ws.row_dimensions[2].height = 20
    ws.merge_cells("A2:B2")
    c = ws["A2"]
    c.value = "GatewayCrypto  |  Prepared by Roman  |  Pre-registration: May 26, 2026"
    c.fill = fill("1A3A5C")
    c.font = font(italic=True, color="AACCCC", size=10)
    c.alignment = align("center")
    c.border = thin_border()

    rows = [
        ("", ""),
        ("EVENT", "next.io — iGaming Expo + Collective"),
        ("GOAL", "Generate qualified pipeline from platforms, operators, and payment partners. Secure intro meetings, qualify fit, and lock concrete next steps before leaving the event."),
        ("GATEWAY CRYPTO", "Crypto payment processing for iGaming — 300+ assets, unified USDT/USDC across chains, zero chargebacks, EUR off-ramp via SEPA/SWIFT, live within 24 hours."),
        ("", ""),
        ("PIPELINE SNAPSHOT", ""),
        ("Total companies engaged", "13"),
        ("Confirmed intro meetings", "6  ✓"),
        ("In-process (to convert or walk-by)", "7"),
    ]

    for i, (label, value) in enumerate(rows, 3):
        ws.row_dimensions[i].height = 30 if label in ("GOAL", "GATEWAY CRYPTO") else 20
        if label in ("EVENT", "GOAL", "GATEWAY CRYPTO", "PIPELINE SNAPSHOT"):
            lc = ws.cell(row=i, column=1, value=label)
            lc.fill = fill(TEAL)
            lc.font = font(bold=True, color=WHITE, size=10)
            lc.alignment = align("center")
            lc.border = thin_border()
            vc = ws.cell(row=i, column=2, value=value)
            vc.fill = fill(MINT)
            vc.font = font(bold=False, color=BLACK)
            vc.alignment = align(wrap=True)
            vc.border = thin_border()
        elif label == "":
            for col in (1, 2):
                c = ws.cell(row=i, column=col, value="")
                c.fill = fill(WHITE)
                c.border = thin_border()
        else:
            lc = ws.cell(row=i, column=1, value=label)
            lc.fill = fill(LGREY)
            lc.font = font(bold=True, color=NAVY)
            lc.alignment = align()
            lc.border = thin_border()
            vc = ws.cell(row=i, column=2, value=value)
            bg = MINT if label == "Confirmed intro meetings" else LGREY
            vc.fill = fill(bg)
            vc.font = font(bold=True if label == "Confirmed intro meetings" else False, color=GREEN if label == "Confirmed intro meetings" else BLACK)
            vc.alignment = align()
            vc.border = thin_border()


# ── Sheet 2: Metrics ──────────────────────────────────────────────────────────

def build_metrics(wb):
    ws = wb.create_sheet("Success Metrics")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [22, 38, 18, 18])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = "Objectives & Success Metrics"
    c.fill = fill(NAVY)
    c.font = Font(bold=True, color=WHITE, size=14, name="Calibri")
    c.alignment = align("center")
    c.border = thin_border()

    sections = [
        ("PRE-EVENT TARGETS", [
            ("Confirmed meetings booked",                   "6/6",      "✓ DONE", GREEN),
            ("In-process converted to meetings",            "Target 4/7 (57%)", "In Progress", TEAL),
            ("Walk-bys planned (stand known)",              "AMG A4  ·  Gaming Ent L18", "Planned", TEAL),
        ]),
        ("AT-EVENT TARGETS", [
            ("Conversation rating collected per meeting",   "Hot / Warm / Cold per company", "At Event", NAVY),
            ("Concrete next step secured per meeting",      "1 per confirmed intro (demo / trial / intro)", "At Event", NAVY),
            ("Objection captured per cold/warm meeting",    "1 specific blocker noted", "At Event", NAVY),
        ]),
        ("POST-EVENT TARGETS (30-DAY)", [
            ("Follow-up message sent",                      "100% of met contacts", "T+48h", TEAL),
            ("Demo or discovery call scheduled",            "4+ meetings", "T+2 weeks", TEAL),
            ("Proposal / commercial discussion opened",     "2+ leads", "T+30 days", TEAL),
            ("New qualified opps added to tracker",         "6+", "T+30 days", TEAL),
        ]),
    ]

    row = 2
    for section_name, items in sections:
        ws.row_dimensions[row].height = 22
        ws.merge_cells(f"A{row}:D{row}")
        c = ws.cell(row=row, column=1, value=section_name)
        c.fill = fill(TEAL)
        c.font = font(bold=True, color=WHITE, size=11)
        c.alignment = align("center")
        c.border = thin_border()
        row += 1

        headers = ["METRIC", "TARGET", "TIMING / STATUS", "STATUS COLOR"]
        for ci, h in enumerate(headers[:3], 1):
            header_cell(ws, row, ci, h, bg="1A3A5C")
        ws.cell(row=row, column=4, value="").fill = fill("1A3A5C")
        ws.cell(row=row, column=4).border = thin_border()
        row += 1

        for i, (metric, target, timing, color) in enumerate(items):
            bg = MINT if i % 2 == 0 else LGREY
            data_cell(ws, row, 1, metric, bg=bg, bold=True, color=NAVY, wrap=True)
            data_cell(ws, row, 2, target, bg=bg, wrap=True)
            data_cell(ws, row, 3, timing, bg=bg, color=color, bold=True, h="center")
            ws.cell(row=row, column=4, value="").fill = fill(f"{color}")
            ws.cell(row=row, column=4).border = thin_border()
            ws.row_dimensions[row].height = 22
            row += 1

        row += 1


# ── Sheet 3: Confirmed Meetings ───────────────────────────────────────────────

def build_confirmed(wb):
    ws = wb.create_sheet("Confirmed Meetings")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [5, 20, 20, 20, 30, 30, 42])
    ws.row_dimensions[1].height = 30
    freeze(ws, "A2")

    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = "Confirmed Intro Meetings (6)"
    c.fill = fill(NAVY)
    c.font = Font(bold=True, color=WHITE, size=14, name="Calibri")
    c.alignment = align("center")
    c.border = thin_border()

    headers = ["#", "COMPANY", "TYPE", "CONTACT", "WEBSITE", "LINKEDIN", "GWC ANGLE"]
    for ci, h in enumerate(headers, 1):
        header_cell(ws, 2, ci, h)

    meetings = [
        (1, "Plaee", "Platform", "Libi Milshtein",
         "https://www.plaee.com/",
         "linkedin.com/company/plaee/ · linkedin.com/in/libi-milshtein/",
         "Platform-level integration — every operator on Plaee inherits crypto automatically. 300+ assets, unified USDT/USDC, zero chargebacks."),
        (2, "Morefin", "PSP Orchestrator", "Miroslav Naydenov",
         "https://www.morefin.com/",
         "linkedin.com/company/morefin/about/ · linkedin.com/in/mironaydenov/",
         "GWC as the crypto rail in their orchestration stack. They route card/APM — we cover the crypto leg. Additive, no overlap."),
        (3, "BigBux", "Operator / Ruffle", "Viktor Atanasovski",
         "https://bigbux.io/",
         "linkedin.com/company/bigbux/about/ · linkedin.com/in/viktoratanasovski/",
         "Direct integration. Ruffle mechanics attract crypto-native users. Irreversible deposits — zero chargeback exposure."),
        (4, "BetHero", "Sportsbook", "Mike Forslund",
         "https://bethero.gg/",
         "linkedin.com/company/hellobethero/ · linkedin.com/in/mikeforslund/",
         "Sports bettors skew crypto-native. Zero chargebacks, instant settlement. Live within 24 hours."),
        (5, "Start2Pay", "PSP", "Tanya Seleznova",
         "https://start2pay.com/",
         "linkedin.com/in/tanya-seleznova-13a04a212/",
         "Partnership play — they cover fiat rails, GWC covers crypto. EUR off-ramp bridges the gap. No competitive overlap."),
        (6, "PayAdmit", "PSP", "Heorhii Kuchuk",
         "https://payadmit.com/",
         "linkedin.com/company/payadmit/ · linkedin.com/in/heorhii-kuchuk-ab5561247/",
         "Same model as Start2Pay. GWC as the crypto rail they don't have natively. Referral or white-label structure."),
    ]

    for i, (num, company, type_, contact, website, linkedin, angle) in enumerate(meetings):
        row = i + 3
        ws.row_dimensions[row].height = 40
        bg = MINT if i % 2 == 0 else LGREY
        data_cell(ws, row, 1, str(num), bg=bg, bold=True, color=WHITE, h="center")
        ws.cell(row=row, column=1).fill = fill(TEAL)
        data_cell(ws, row, 2, company, bg=bg, bold=True, color=NAVY)
        data_cell(ws, row, 3, type_, bg=bg, color=TEAL)
        data_cell(ws, row, 4, contact, bg=bg)
        data_cell(ws, row, 5, website, bg=bg, color=TEAL)
        data_cell(ws, row, 6, linkedin, bg=bg, color=TEAL, wrap=True)
        data_cell(ws, row, 7, angle, bg=bg, wrap=True)


# ── Sheet 4: In-Process Pipeline ─────────────────────────────────────────────

def build_pipeline(wb):
    ws = wb.create_sheet("In-Process Pipeline")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [20, 22, 26, 8, 26, 30, 20])
    ws.row_dimensions[1].height = 30
    freeze(ws, "A2")

    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = "In-Process Pipeline (7)  —  Chase / Walk-By"
    c.fill = fill(NAVY)
    c.font = Font(bold=True, color=WHITE, size=14, name="Calibri")
    c.alignment = align("center")
    c.border = thin_border()

    headers = ["COMPANY", "TYPE", "CONTACT", "STAND", "WEBSITE", "LINKEDIN", "ACTION REQUIRED"]
    for ci, h in enumerate(headers, 1):
        header_cell(ws, 2, ci, h)

    companies = [
        ("AMG Platform",    "Platform",             "Mark Abdilla",                 "A4",  "https://www.amgaminggroup.com/",   "linkedin.com/company/amgplatform/ · linkedin.com/in/markabdilla/",          "Walk-by confirmed — approach on floor"),
        ("Gaming Ent",      "Platform",             "Boyko Boev",                   "L18", "https://gaming-ent.com/",          "linkedin.com/company/gaming-ent/ · linkedin.com/in/boyko-boev-4b098466/",   "Walk-by confirmed — approach on floor"),
        ("Kanggiten",       "Platform",             "Sergey Shibkih",               "—",   "https://kanggiten.com/",           "linkedin.com/company/kanggiten/ · linkedin.com/in/sergey-shibkih/",          "Chase pre-event"),
        ("Slikair",         "Payment Orchestrator", "Tzach Toporek",                "—",   "https://www.slikair.com/",         "linkedin.com/company/slikair/ · linkedin.com/in/tzach-toporek/",             "Chase pre-event"),
        ("Comtrade Gaming", "Platform",             "Steven Valentine / Ales Gornjec", "—","https://www.comtradegaming.com/",  "linkedin.com/in/steven-valentine-021a0014/ · linkedin.com/in/alesgornjec/",  "Chase pre-event — 2 contacts"),
        ("Bejoynd",         "Platform",             "Fredrik Cedell",               "—",   "https://www.bejoynd.com/",         "linkedin.com/company/bejoynd/ · linkedin.com/in/fredrik-cedell-765b105b/",   "Chase pre-event"),
        ("Ace Systems",     "Platform",             "William Lövqvist",             "—",   "https://acesystem.io/",            "linkedin.com/company/acesystem/about/ · linkedin.com/in/william-lovqvist/",  "Chase pre-event"),
    ]

    for i, (company, type_, contact, stand, website, linkedin, action) in enumerate(companies):
        row = i + 3
        ws.row_dimensions[row].height = 35
        bg = MINT if i % 2 == 0 else LGREY
        data_cell(ws, row, 1, company, bg=bg, bold=True, color=NAVY)
        data_cell(ws, row, 2, type_, bg=bg, color=TEAL)
        data_cell(ws, row, 3, contact, bg=bg)
        stand_color = GREEN if stand != "—" else BLACK
        stand_bold = stand != "—"
        data_cell(ws, row, 4, stand, bg=bg, bold=stand_bold, color=stand_color, h="center")
        data_cell(ws, row, 5, website, bg=bg, color=TEAL)
        data_cell(ws, row, 6, linkedin, bg=bg, color=TEAL, wrap=True)
        action_color = GREEN if "Walk-by" in action else TEAL
        data_cell(ws, row, 7, action, bg=bg, bold=True, color=action_color)


# ── Sheet 5: Event Schedule ───────────────────────────────────────────────────

def build_schedule(wb):
    ws = wb.create_sheet("Event Schedule")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [16, 42, 14, 28])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = "Event Schedule"
    c.fill = fill(NAVY)
    c.font = Font(bold=True, color=WHITE, size=14, name="Calibri")
    c.alignment = align("center")
    c.border = thin_border()

    headers = ["DATE / TIME", "ACTIVITY", "PRIORITY", "NOTES"]
    for ci, h in enumerate(headers, 1):
        header_cell(ws, 2, ci, h)

    priority_colors = {"HIGH": GREEN, "MEDIUM": TEAL, "TBC": "888888"}
    schedule = [
        ("May 26",    "Pre-registration event",                           "HIGH",   "Arrive early — first impressions matter"),
        ("TBC",       "Rooftop side event #1",                            "HIGH",   "Update with time / venue"),
        ("TBC",       "Rooftop side event #2",                            "HIGH",   "Update with time / venue"),
        ("TBC",       "Networking event",                                  "MEDIUM", "Update with time / venue"),
        ("Expo days", "6 confirmed intro meetings",                        "HIGH",   "Prioritise morning slots — energy is highest early"),
        ("Expo days", "Walk-by: AMG Platform — Stand A4",                 "HIGH",   "Approach if pre-event intro not locked"),
        ("Expo days", "Walk-by: Gaming Ent — Stand L18",                  "HIGH",   "Approach if pre-event intro not locked"),
    ]

    for i, (date, activity, priority, notes) in enumerate(schedule):
        row = i + 3
        ws.row_dimensions[row].height = 22
        bg = MINT if i % 2 == 0 else LGREY
        data_cell(ws, row, 1, date, bg=bg, bold=True, color=NAVY, h="center")
        data_cell(ws, row, 2, activity, bg=bg)
        data_cell(ws, row, 3, priority, bg=bg, bold=True, color=priority_colors.get(priority, BLACK), h="center")
        data_cell(ws, row, 4, notes, bg=bg, wrap=True)


# ── Sheet 6: Post-Event Actions ───────────────────────────────────────────────

def build_actions(wb):
    ws = wb.create_sheet("Post-Event Actions")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [12, 46, 12, 20, 12])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = "Post-Event Action Plan"
    c.fill = fill(NAVY)
    c.font = Font(bold=True, color=WHITE, size=14, name="Calibri")
    c.alignment = align("center")
    c.border = thin_border()

    headers = ["TIMELINE", "ACTION", "OWNER", "WORKFLOW / TOOL", "DONE?"]
    for ci, h in enumerate(headers, 1):
        header_cell(ws, 2, ci, h)

    actions = [
        ("T+48h",      "Send follow-up LinkedIn message to all 6 confirmed meeting contacts",          "Roman", "reply_drafting.md",          "☐"),
        ("T+48h",      "Send outreach to in-process contacts not converted pre-event",                 "Roman", "linkedin_outreach.md",        "☐"),
        ("T+2 weeks",  "Book demos / discovery calls for all Hot-rated conversations",                 "Roman", "—",                          "☐"),
        ("T+2 weeks",  "Update outreach tracker with outcome for every company touched at the event",  "Roman", "outreach_tracker.md",         "☐"),
        ("T+30 days",  "Open proposal or commercial discussion for 2+ qualified leads",                "Roman", "—",                          "☐"),
        ("T+30 days",  "Review in-process pipeline — move or close stalled conversations",             "Roman", "telegram_followup_batch.md",  "☐"),
        ("T+30 days",  "Post-event debrief: what worked, what didn't — update workflows if needed",    "Roman", "CLAUDE.md / workflows/",      "☐"),
    ]

    timeline_colors = {"T+48h": RED, "T+2 weeks": TEAL, "T+30 days": NAVY}
    for i, (timeline, action, owner, workflow, done) in enumerate(actions):
        row = i + 3
        ws.row_dimensions[row].height = 28
        bg = MINT if i % 2 == 0 else LGREY
        data_cell(ws, row, 1, timeline, bg=bg, bold=True, color=timeline_colors.get(timeline, BLACK), h="center")
        data_cell(ws, row, 2, action, bg=bg, wrap=True)
        data_cell(ws, row, 3, owner, bg=bg, h="center")
        data_cell(ws, row, 4, workflow, bg=bg, color=TEAL)
        data_cell(ws, row, 5, done, bg=bg, h="center", bold=True)


# ── Main ──────────────────────────────────────────────────────────────────────

def build():
    wb = openpyxl.Workbook()
    build_overview(wb)
    build_metrics(wb)
    build_confirmed(wb)
    build_pipeline(wb)
    build_schedule(wb)
    build_actions(wb)

    out = "/Users/roman/Downloads/nextio_2026_Event_Strategy_Plan.xlsx"
    wb.save(out)
    print(f"Saved → {out}")


build()
