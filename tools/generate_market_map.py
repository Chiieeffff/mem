"""
Market intelligence map generator for GatewayCrypto.
Researches 100 B2B iGaming platforms: payment processors used + company size.
Outputs XLSX with blank manual-entry columns for contacts, outreach, and status.

Usage:
  # Research + export in one shot
  python tools/generate_market_map.py --leads '[
    {"company": "EveryMatrix", "website": "https://everymatrix.com"},
    {"company": "SOFTSWISS", "website": "https://softswiss.com"}
  ]' --output .tmp/market_map.xlsx

  # Research only (dump JSON for manual review before export)
  python tools/generate_market_map.py --leads '[...]' --research-only

  # Export from previously saved research JSON
  python tools/generate_market_map.py --from-json .tmp/research.json --output .tmp/market_map.xlsx
"""

import argparse
import json
import os
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from dotenv import load_dotenv

load_dotenv()

# ── colours (match existing export_prospects.py palette) ─────────────────────
NAVY  = "0D1B2A"
TEAL  = "1A7A6A"
MINT  = "E8F5F2"
WHITE = "FFFFFF"
BLACK = "1A1A1A"
LGREY = "F4F4F4"
MGREY = "CCCCCC"
AMBER = "FFF3CD"  # highlight for manual-entry columns

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color=BLACK, size=10, italic=False):
    return Font(bold=bold, color=color, size=size, name="Calibri", italic=italic)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _border():
    s = Side(style="thin", color=MGREY)
    return Border(left=s, right=s, top=s, bottom=s)


# ── column schema ─────────────────────────────────────────────────────────────
# (header, col_width, field_key, manual_entry)
COLUMNS = [
    ("Company",           28, "company",          False),
    ("Website",           28, "website",          False),
    ("HQ / Region",       22, "region",           False),
    ("Est. Size",         18, "size",             False),
    ("Founded",           14, "founded",          False),
    ("Crypto Processor",  35, "crypto_processor", False),
    ("Fiat Processor",    35, "fiat_processor",   False),
    ("Who We Know",       28, "contact",          True),
    ("Talked To",         28, "talked_to",        True),
    ("Status",            18, "status",           True),
]


# ── research ──────────────────────────────────────────────────────────────────

def search_company(company: str) -> list[dict]:
    """Run 3 targeted Tavily queries focused on processor + size intelligence."""
    from tavily import TavilyClient

    api_key = os.getenv("TAVILY_API_KEY")
    if not api_key:
        raise RuntimeError("TAVILY_API_KEY not set in .env")

    client = TavilyClient(api_key=api_key)

    queries = [
        f'"{company}" iGaming crypto payment processor CoinsPaid NowPayments Utorg GatewayCrypto BitcoinPay CryptoPay',
        f'"{company}" iGaming fiat payment processor Nuvei Paysafe Worldpay Adyen SafeCharge Trustly Skrill PaySafeCard',
        f'"{company}" iGaming employees headcount revenue funding founded year established',
        f'"{company}" iGaming platform integrations payment partners OR payment methods',
    ]

    results = []
    for query in queries:
        try:
            response = client.search(
                query=query,
                max_results=3,
                search_depth="basic",
                include_raw_content=False,
            )
            hits = [
                {
                    "title": r.get("title", ""),
                    "url": r.get("url", ""),
                    "content": r.get("content", ""),
                }
                for r in response.get("results", [])
            ]
            results.append({"query": query, "results": hits})
        except Exception as e:
            results.append({"query": query, "error": str(e), "results": []})

    return results


def scrape_company(website: str) -> dict:
    """Scrape integrations/partners pages for payment processor mentions."""
    import requests
    from bs4 import BeautifulSoup
    from urllib.parse import urljoin, urlparse
    import re

    CANDIDATE_PATHS = ["/", "/integrations", "/partners", "/technology", "/platform", "/payments"]
    HEADERS = {"User-Agent": "Mozilla/5.0 (compatible; research-bot/1.0)"}
    TIMEOUT = 10

    parsed = urlparse(website)
    origin = f"{parsed.scheme}://{parsed.netloc}"
    pages = []

    for path in CANDIDATE_PATHS:
        url = urljoin(origin, path)
        try:
            resp = requests.get(url, headers=HEADERS, timeout=TIMEOUT)
            if resp.status_code != 200:
                continue
            soup = BeautifulSoup(resp.text, "html.parser")
            for tag in soup(["script", "style", "nav", "footer", "head"]):
                tag.decompose()
            text = soup.get_text(separator=" ", strip=True)
            text = re.sub(r"\s{3,}", "  ", text)
            pages.append({"path": path, "text": text[:3000]})
        except Exception:
            continue

    return {"url": website, "pages": pages}


def research_lead(lead: dict) -> dict:
    company = lead.get("company", "").strip()
    website = (lead.get("website") or "").strip() or None

    result = {
        "company": company,
        "website": website or "",
        "region": lead.get("region", ""),
        "search_results": [],
        "scrape_results": None,
        "errors": [],
    }

    try:
        result["search_results"] = search_company(company)
    except Exception as e:
        result["errors"].append(f"Search failed: {e}")

    if website:
        try:
            result["scrape_results"] = scrape_company(website)
        except Exception as e:
            result["errors"].append(f"Scrape failed: {e}")

    return result


def run_research(leads: list[dict], workers: int = 2) -> list[dict]:
    results = [None] * len(leads)
    with ThreadPoolExecutor(max_workers=workers) as executor:
        future_to_index = {
            executor.submit(research_lead, lead): i for i, lead in enumerate(leads)
        }
        for future in as_completed(future_to_index):
            idx = future_to_index[future]
            try:
                results[idx] = future.result()
            except Exception as e:
                results[idx] = {
                    "company": leads[idx].get("company", "unknown"),
                    "website": leads[idx].get("website", ""),
                    "region": leads[idx].get("region", ""),
                    "search_results": [],
                    "scrape_results": None,
                    "errors": [str(e)],
                }
    return results


# ── XLSX export ───────────────────────────────────────────────────────────────

def build_xlsx(rows: list[dict], output_path: str) -> None:
    """
    rows: list of dicts with keys matching COLUMNS field_keys.
    Manual columns (contact, talked_to, status, notes) may be absent or empty.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Market Map"

    ncols = len(COLUMNS)

    # title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    title = ws.cell(row=1, column=1, value="iGaming B2B Platform Market Map — GatewayCrypto")
    title.fill = _fill(NAVY)
    title.font = Font(bold=True, color=WHITE, size=13, name="Calibri")
    title.alignment = _align(h="center")
    ws.row_dimensions[1].height = 28

    # date row
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    date_cell = ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  {len(rows)} platforms")
    date_cell.fill = _fill(TEAL)
    date_cell.font = _font(color=WHITE, size=9, italic=True)
    date_cell.alignment = _align(h="right")

    # instruction row
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=ncols)
    note = ws.cell(row=3, column=1, value="Columns shaded amber are for manual entry: Who We Know / Talked To / Status")
    note.fill = _fill(AMBER)
    note.font = _font(size=9, italic=True)
    note.alignment = _align(h="left")

    # header row
    for col_idx, (header, width, _, manual) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.fill = _fill(AMBER) if manual else _fill(TEAL)
        cell.font = _font(bold=True, color=BLACK if manual else WHITE, size=10)
        cell.alignment = _align(h="center")
        cell.border = _border()
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
    ws.row_dimensions[4].height = 22

    # data rows
    for row_idx, row_data in enumerate(rows, start=5):
        stripe = row_idx % 2 == 0
        for col_idx, (_, _, field, manual) in enumerate(COLUMNS, start=1):
            value = row_data.get(field, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if manual:
                cell.fill = _fill(AMBER if stripe else "FFFBF0")
            else:
                cell.fill = _fill(MINT if stripe else WHITE)
            cell.font = _font(size=10)
            cell.alignment = _align(wrap=True)
            cell.border = _border()
        ws.row_dimensions[row_idx].height = 18

    ws.freeze_panes = "A5"
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)


# ── main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Generate iGaming market intelligence map")
    parser.add_argument("--leads", help="JSON array of {company, website, region} objects")
    parser.add_argument("--from-json", help="Path to previously saved research JSON")
    parser.add_argument("--rows-json", help="JSON array of pre-synthesized row dicts to export directly (skips research)")
    parser.add_argument("--workers", type=int, default=2, help="Parallel workers (default: 2 for Tavily rate limits)")
    parser.add_argument("--output", default=None, help="Output XLSX path")
    parser.add_argument("--research-only", action="store_true", help="Dump research JSON to stdout, skip XLSX")

    args = parser.parse_args()

    if not any([args.leads, args.from_json, args.rows_json]):
        parser.error("one of --leads, --from-json, or --rows-json is required")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = args.output or os.path.abspath(f".tmp/market_map_{timestamp}.xlsx")

    # ── mode 1: export from pre-synthesized rows ──────────────────────────────
    if args.rows_json:
        try:
            rows = json.loads(args.rows_json)
        except json.JSONDecodeError as e:
            print(f"ERROR: --rows-json is not valid JSON: {e}", file=sys.stderr)
            sys.exit(1)
        build_xlsx(rows, output_path)
        print(output_path)
        return

    # ── mode 2: load from previously saved research JSON ─────────────────────
    if args.from_json:
        with open(args.from_json, "r", encoding="utf-8") as f:
            research = json.load(f)
        print(json.dumps(research, indent=2, ensure_ascii=False))
        return

    # ── mode 3: run research from leads list ──────────────────────────────────
    try:
        leads = json.loads(args.leads)
    except json.JSONDecodeError as e:
        print(f"ERROR: --leads is not valid JSON: {e}", file=sys.stderr)
        sys.exit(1)

    if not isinstance(leads, list) or not leads:
        print("ERROR: --leads must be a non-empty JSON array", file=sys.stderr)
        sys.exit(1)

    research = run_research(leads, workers=args.workers)

    if args.research_only:
        print(json.dumps(research, indent=2, ensure_ascii=False))
        return

    # With --leads and no --research-only: dump JSON for Claude to synthesize
    # (Claude reads search results and constructs the rows dict, then calls --rows-json)
    print(json.dumps(research, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
