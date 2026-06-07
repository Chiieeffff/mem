from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "next.io 2026"

# ── Palette ──────────────────────────────────────────────────────────────────
NAVY       = "1C2B3A"
GOLD       = "C9A84C"
GREEN      = "1A7A4A"
GREEN_HDR  = "155F3A"
AMBER      = "B45309"
AMBER_HDR  = "92400E"
LIGHT_NAVY = "2E3F52"
BG_GREEN   = "EAF5EE"
BG_AMBER   = "FEF3C7"
BG_GREY    = "F0F2F5"
WHITE      = "FFFFFF"
TEXT_DARK  = "1C2B3A"
TEXT_MID   = "4A5568"
LINK_CLR   = "1D4ED8"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def fnt(bold=False, size=9, color=TEXT_DARK, italic=False, underline=None):
    return Font(name="Calibri", bold=bold, size=size, color=color,
                italic=italic, underline=underline)

def aln(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def bx(color="CCCCCC"):
    s = Side(style="thin", color=color)
    return Border(top=s, bottom=s, left=s, right=s)

def sc(ws, r, c, val, bold=False, size=9, color=TEXT_DARK,
       bg=None, h="left", v="center", wrap=False, italic=False, border=None):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font  = fnt(bold=bold, size=size, color=color, italic=italic)
    cell.alignment = aln(h=h, v=v, wrap=wrap)
    if bg:     cell.fill   = fill(bg)
    if border: cell.border = border
    return cell

def mg(ws, r1, c1, r2, c2, val, bold=False, size=9, color=TEXT_DARK,
       bg=None, h="left", v="center", wrap=False, italic=False, border=None):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = ws.cell(row=r1, column=c1, value=val)
    cell.font  = fnt(bold=bold, size=size, color=color, italic=italic)
    cell.alignment = aln(h=h, v=v, wrap=wrap)
    if bg:     cell.fill   = fill(bg)
    if border: cell.border = border
    return cell

def hl(ws, r, c, url, display, bg=None, v="center"):
    cell = ws.cell(row=r, column=c, value=display)
    cell.hyperlink  = url
    cell.font       = Font(name="Calibri", size=8, color=LINK_CLR, underline="single")
    cell.alignment  = aln(h="left", v=v)
    cell.border     = bx()
    if bg: cell.fill = fill(bg)
    return cell

def mg_hl(ws, r1, c1, r2, c2, url, display, bg=None):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = ws.cell(row=r1, column=c1, value=display)
    cell.hyperlink  = url
    cell.font       = Font(name="Calibri", size=8, color=LINK_CLR, underline="single")
    cell.alignment  = aln(h="left", v="center")
    cell.border     = bx()
    if bg: cell.fill = fill(bg)
    return cell

# ── Column widths ──────────────────────────────────────────────────────────
# 1=margin | 2=company | 3=website | 4=co.LN | 5=contact | 6=stand | 7=direction | 8=status
# 9=gap
# 10=company | 11=website | 12=co.LN | 13=contact | 14=stand | 15=direction | 16=status
widths = {
    1:1.5, 2:17, 3:16, 4:16, 5:20, 6:6, 7:34, 8:10,
    9:2,
    10:17, 11:16, 12:16, 13:20, 14:6, 15:34, 16:10,
}
for c, w in widths.items():
    ws.column_dimensions[get_column_letter(c)].width = w

# ── Row heights ────────────────────────────────────────────────────────────
ws.row_dimensions[1].height = 3
ws.row_dimensions[2].height = 28
ws.row_dimensions[3].height = 15
ws.row_dimensions[4].height = 13
for r in range(5, 20):
    # odd rows = name row, even rows = type/url row
    ws.row_dimensions[r].height = 15 if (r % 2 != 0) else 12
ws.row_dimensions[19].height = 13   # footer

# ── ROW 1: margin strip ────────────────────────────────────────────────────
for c in range(1, 17):
    ws.cell(row=1, column=c).fill = fill(NAVY)

# ── ROW 2: title ──────────────────────────────────────────────────────────
for c in range(1, 17):
    ws.cell(row=2, column=c).fill = fill(NAVY)
mg(ws, 2, 1, 2, 16,
   "next.io 2026  ·  iGaming Expo + Collective  ·  Valletta  |  GatewayCrypto Outreach",
   bold=True, size=13, color=WHITE, bg=NAVY, h="center", v="center")

# ── ROW 3: group headers ───────────────────────────────────────────────────
for c in range(1, 17):
    ws.cell(row=3, column=c).fill = fill(LIGHT_NAVY)

mg(ws, 3, 2, 3, 7, "  SCHEDULED MEETINGS  (6)",
   bold=True, size=9, color=WHITE, bg=GREEN, h="left", v="center")
sc(ws, 3, 8, "Confirmed", bold=True, size=8, color=WHITE, bg=GREEN, h="center")

ws.cell(row=3, column=9).fill = fill(NAVY)

mg(ws, 3, 10, 3, 15, "  IN PROCESS  (7)  —  Chase / Walk-By",
   bold=True, size=9, color=WHITE, bg=AMBER, h="left", v="center")
sc(ws, 3, 16, "Status", bold=True, size=8, color=WHITE, bg=AMBER, h="center")

# ── ROW 4: sub-headers ────────────────────────────────────────────────────
sub = ["Company / Type", "Website", "Company LinkedIn", "Contact / LinkedIn",
       "Stand", "Discussion Direction", ""]
l_cols = list(range(2, 9))
r_cols = list(range(10, 17))

for c, h in zip(l_cols, sub):
    sc(ws, 4, c, h, bold=True, size=8, color=GREEN_HDR,
       bg=BG_GREEN, h="center", v="center", border=bx("A7D4B8"))

ws.cell(row=4, column=9).fill = fill(WHITE)

for c, h in zip(r_cols, sub):
    sc(ws, 4, c, h, bold=True, size=8, color=AMBER_HDR,
       bg=BG_AMBER, h="center", v="center", border=bx("E8C97A"))

# ── Data ──────────────────────────────────────────────────────────────────
# Each tuple: (name, type, web_url, web_txt, cln_url, cln_txt,
#              contact_name, pln_url, pln_txt, stand, direction)

confirmed = [
    ("PLAEE", "Platform",
     "https://www.plaee.com/", "plaee.com",
     "https://www.linkedin.com/company/plaee/", "↗ /plaee",
     "Libi Milshtein",
     "https://www.linkedin.com/in/libi-milshtein/", "↗ /libi-milshtein",
     "C8",
     "Platform integration — every operator on Plaee inherits crypto by default; one deal, full operator base unlocked"),

    ("MOREFIN", "PSP Orchestrator",
     "https://www.morefin.com/", "morefin.com",
     "https://www.linkedin.com/company/morefin/about/", "↗ /morefin",
     "Miroslav Naydenov",
     "https://www.linkedin.com/in/mironaydenov/", "↗ /mironaydenov",
     "",
     "Partnership — GWC as crypto rail inside their stack; they keep card/APM, we add the crypto leg; additive, zero overlap"),

    ("BIGBUX", "Operator / Ruffle",
     "https://bigbux.io/", "bigbux.io",
     "https://www.linkedin.com/company/bigbux/about/", "↗ /bigbux",
     "Viktor Atanasovski",
     "https://www.linkedin.com/in/viktoratanasovski/", "↗ /viktoratanasovski",
     "L13",
     "Direct integration — ruffle mechanics attract crypto-native users; crypto deposits are irreversible, zero chargeback exposure"),

    ("BETHERO", "Sportsbook",
     "https://bethero.gg/", "bethero.gg",
     "https://www.linkedin.com/company/hellobethero/", "↗ /hellobethero",
     "Mike Forslund",
     "https://www.linkedin.com/in/mikeforslund/", "↗ /mikeforslund",
     "",
     "Direct integration — sports bettors skew crypto-native; instant settlement, 300+ assets, live within 24h"),

    ("START2PAY", "PSP",
     "https://start2pay.com/", "start2pay.com",
     "https://www.linkedin.com/company/start2pay/about/", "↗ /start2pay",
     "Tanya Seleznova",
     "https://www.linkedin.com/in/tanya-seleznova-13a04a212/", "↗ /tanya-seleznova",
     "",
     "Partnership — they own fiat rails, GWC owns crypto leg; EUR off-ramp via SEPA/SWIFT bridges shared merchants"),

    ("PAYADMIT", "PSP",
     "https://payadmit.com/", "payadmit.com",
     "https://www.linkedin.com/company/payadmit/", "↗ /payadmit",
     "Heorhii Kuchuk",
     "https://www.linkedin.com/in/heorhii-kuchuk-ab5561247/", "↗ /heorhii-kuchuk",
     "",
     "Partnership — referral or white-label structure; GWC as native crypto rail; zero competition on the fiat side"),
]

# (name, type, web_url, web_txt, cln_url, cln_txt,
#  contact_name, pln_url, pln_txt, stand, direction, status)
in_process = [
    ("AMG PLATFORM", "Platform",
     "https://www.amgaminggroup.com/", "amgaminggroup.com",
     "https://www.linkedin.com/company/amgplatform/", "↗ /amgplatform",
     "Mark Abdilla",
     "https://www.linkedin.com/in/markabdilla/", "↗ /markabdilla",
     "A4",
     "Platform integration — stand A4 confirmed; approach on floor if pre-event intro not locked",
     "Walk-By ✓"),

    ("GAMING ENT", "Platform",
     "https://gaming-ent.com/", "gaming-ent.com",
     "https://www.linkedin.com/company/gaming-ent/", "↗ /gaming-ent",
     "Boyko Boev",
     "https://www.linkedin.com/in/boyko-boev-4b098466/", "↗ /boyko-boev",
     "L18",
     "Platform integration — stand L18 confirmed; operators get crypto by default, same leverage as Plaee",
     "Walk-By ✓"),

    ("KANGGITEN", "Platform",
     "https://kanggiten.com/", "kanggiten.com",
     "https://www.linkedin.com/company/kanggiten/", "↗ /kanggiten",
     "Sergey Shibkih",
     "https://www.linkedin.com/in/sergey-shibkih/", "↗ /sergey-shibkih",
     "",
     "Platform integration — operators inherit crypto automatically; chase pre-event for stand",
     "Chase"),

    ("SLIKAIR", "Payment Orchestrator",
     "https://www.slikair.com/", "slikair.com",
     "https://www.linkedin.com/company/slikair/", "↗ /slikair",
     "Tzach Toporek",
     "https://www.linkedin.com/in/tzach-toporek/", "↗ /tzach-toporek",
     "",
     "Partnership — GWC as crypto rail in their orchestration stack; mirrors Morefin model; chase pre-event",
     "Chase"),

    ("COMTRADE GAMING", "Platform",
     "https://www.comtradegaming.com/", "comtradegaming.com",
     "https://www.linkedin.com/company/comtrade-gaming/", "↗ /comtrade-gaming",
     "Steven Valentine / Ales G.",
     "https://www.linkedin.com/in/steven-valentine-021a0014/", "↗ /steven-valentine",
     "",
     "Platform integration — lead Steven (commercial), Ales (technical); split conversation if both present",
     "Chase (×2)"),

    ("BEJOYND", "Platform",
     "https://www.bejoynd.com/", "bejoynd.com",
     "https://www.linkedin.com/company/bejoynd/", "↗ /bejoynd",
     "Fredrik Cedell",
     "https://www.linkedin.com/in/fredrik-cedell-765b105b/", "↗ /fredrik-cedell",
     "",
     "Platform integration — operators inherit crypto automatically; chase pre-event for stand",
     "Chase"),

    ("ACE SYSTEMS", "Platform",
     "https://acesystem.io/", "acesystem.io",
     "https://www.linkedin.com/company/acesystem/about/", "↗ /acesystem",
     "William Lövqvist",
     "https://www.linkedin.com/in/william-lovqvist/", "↗ /william-lovqvist",
     "",
     "Platform integration — operators inherit crypto automatically; chase pre-event for stand",
     "Chase"),
]

# ── Render left panel (confirmed meetings) — rows 5-16 ────────────────────
S = 5
for i, d in enumerate(confirmed):
    (name, ctype, web_url, web_txt, cln_url, cln_txt,
     contact, pln_url, pln_txt, stand, direction) = d
    r   = S + i * 2
    bg  = WHITE if i % 2 == 0 else BG_GREEN

    # Company name (r,2) — bold top-aligned
    sc(ws, r,   2, name,    bold=True,  size=9, color=TEXT_DARK, bg=bg, v="bottom", border=bx())
    # Company type (r+1,2) — italic
    sc(ws, r+1, 2, ctype,   italic=True, size=8, color=TEXT_MID,  bg=bg, v="top",    border=bx())

    # Website merged r:r+1
    mg_hl(ws, r, 3, r+1, 3, web_url, web_txt, bg=bg)
    # Company LN merged r:r+1
    mg_hl(ws, r, 4, r+1, 4, cln_url, cln_txt, bg=bg)

    # Contact name (r,5)
    sc(ws, r,   5, contact, size=9,  color=TEXT_DARK, bg=bg, v="bottom", border=bx())
    # Contact LN (r+1,5)
    hl(ws, r+1, 5, pln_url, pln_txt, bg=bg, v="top")

    # Stand merged r:r+1
    ws.merge_cells(start_row=r, start_column=6, end_row=r+1, end_column=6)
    sc(ws, r, 6, stand, bold=True, size=9, color=TEXT_DARK, bg=bg,
       h="center", v="center", border=bx())

    # Direction merged r:r+1
    ws.merge_cells(start_row=r, start_column=7, end_row=r+1, end_column=7)
    sc(ws, r, 7, direction, italic=True, size=8, color=TEXT_DARK, bg=bg,
       v="center", wrap=True, border=bx())

    # Confirmed ✓ merged r:r+1
    ws.merge_cells(start_row=r, start_column=8, end_row=r+1, end_column=8)
    sc(ws, r, 8, "✓", bold=True, size=12, color=GREEN, bg=bg,
       h="center", v="center", border=bx())

# ── Render right panel (in-process) — rows 5-18 ───────────────────────────
for i, d in enumerate(in_process):
    (name, ctype, web_url, web_txt, cln_url, cln_txt,
     contact, pln_url, pln_txt, stand, direction, status) = d
    r   = S + i * 2
    bg  = WHITE if i % 2 == 0 else BG_AMBER

    sc(ws, r,   10, name,    bold=True,  size=9, color=TEXT_DARK, bg=bg, v="bottom", border=bx())
    sc(ws, r+1, 10, ctype,   italic=True, size=8, color=TEXT_MID,  bg=bg, v="top",    border=bx())

    mg_hl(ws, r, 11, r+1, 11, web_url, web_txt, bg=bg)
    mg_hl(ws, r, 12, r+1, 12, cln_url, cln_txt, bg=bg)

    sc(ws, r,   13, contact, size=9,  color=TEXT_DARK, bg=bg, v="bottom", border=bx())
    hl(ws, r+1, 13, pln_url, pln_txt, bg=bg, v="top")

    ws.merge_cells(start_row=r, start_column=14, end_row=r+1, end_column=14)
    sc(ws, r, 14, stand, bold=True, size=9, color=TEXT_DARK, bg=bg,
       h="center", v="center", border=bx())

    ws.merge_cells(start_row=r, start_column=15, end_row=r+1, end_column=15)
    sc(ws, r, 15, direction, italic=True, size=8, color=TEXT_DARK, bg=bg,
       v="center", wrap=True, border=bx())

    ws.merge_cells(start_row=r, start_column=16, end_row=r+1, end_column=16)
    stat_color = GREEN if "Walk" in status else AMBER
    sc(ws, r, 16, status, bold=True, size=8, color=stat_color, bg=bg,
       h="center", v="center", border=bx())

# ── Fill separator + left overflow (rows 17-18 left panel has no data) ─────
for r in range(5, 19):
    ws.cell(row=r, column=9).fill = fill(WHITE)

for r in [17, 18]:
    for c in range(2, 9):
        ws.cell(row=r, column=c).fill = fill(BG_GREY)

# ── Footer row ─────────────────────────────────────────────────────────────
FR = 19
for c in range(1, 17):
    ws.cell(row=FR, column=c).fill = fill(NAVY)
mg(ws, FR, 1, FR, 16,
   "Pre-registration: May 26, 2026  ·  Prepared by Roman  ·  GatewayCrypto"
   "  |  300+ assets · USDT/USDC · zero chargebacks · EUR off-ramp · live in 24h",
   italic=True, size=8, color=GOLD, bg=NAVY, h="center", v="center")

# ── Sheet settings ─────────────────────────────────────────────────────────
ws.sheet_view.showGridLines = False
ws.page_setup.orientation   = "landscape"
ws.page_setup.fitToWidth    = 1
ws.page_setup.fitToHeight   = 1
ws.print_area = "A1:P19"

out = "/Users/roman/claude/next_io_2026_event_plan.xlsx"
wb.save(out)
print(f"Saved: {out}")
